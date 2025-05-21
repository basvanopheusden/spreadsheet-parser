import csv
import re
from pathlib import Path
from typing import Dict, Iterable, List, Union, Optional
import logging

from .models import Company, CANONICAL_HEADERS

logger = logging.getLogger(__name__)


def _sanitize(text: str) -> str:
    """Normalize header names for matching."""
    return "".join(ch.lower() for ch in text if ch.isalnum())


_EXCLUDE_PATTERNS = [
    r"\binstitute\b",
    r"\buniversity\b",
    r"\bcollege\b",
    r"\bfoundation\b",
    r"\bcentre\b",
    r"\bcenter\b",
]


_MONTH_MAP = {
    "Jan": "1",
    "Feb": "2",
    "Mar": "3",
    "Apr": "4",
    "May": "5",
    "Jun": "6",
    "Jul": "7",
    "Aug": "8",
    "Sep": "9",
    "Oct": "10",
    "Nov": "11",
    "Dec": "12",
}


def _decode_lines(file_obj) -> Iterable[str]:
    """Yield decoded lines from a binary file object with fallback."""

    def decode(b: bytes, *, first: bool = False) -> str:
        enc = "utf-8-sig" if first else "utf-8"
        try:
            return b.decode(enc)
        except UnicodeDecodeError:
            return b.decode("latin-1").encode("utf-8", "replace").decode("utf-8")

    first = True
    for bline in file_obj:
        yield decode(bline, first=first)
        first = False


def _fix_employee_range(text: str) -> str:
    """Normalize malformed employee range strings."""

    m = re.match(r"^(\d+)-([A-Za-z]{3})$", text)
    if m:
        day, mon = m.groups()
        mon_num = _MONTH_MAP.get(mon.title())
        if mon_num:
            return f"{mon_num}-{day}"

    m = re.match(r"^([A-Za-z]{3})-(\d+)$", text)
    if m:
        mon, year = m.groups()
        mon_num = _MONTH_MAP.get(mon.title())
        if mon_num:
            return f"{mon_num}-{year}"

    return text


def _employee_bounds(text: str) -> tuple[Optional[int], Optional[int]]:
    """Return the numeric bounds described by an employee range string."""

    if text is None:
        return (None, None)

    text = str(text).replace(",", "").strip()
    m = re.match(r"^(\d+)\s*[-–]\s*(\d+)$", text)
    if m:
        return (int(m.group(1)), int(m.group(2)))

    m = re.match(r"^(\d+)\s*\+$", text)
    if m:
        return (int(m.group(1)), None)

    digits = re.findall(r"\d+", text)
    if digits:
        n = int(digits[0])
        return (n, n)

    return (None, None)


def _employee_range_contains(default: str, actual: str) -> bool:
    """Return True if ``actual`` falls within ``default`` bounds."""

    d_min, d_max = _employee_bounds(default)
    a_min, a_max = _employee_bounds(actual)

    if d_min is None or a_min is None:
        return False

    if a_min < d_min:
        return False

    if d_max is not None:
        if a_max is None or a_max > d_max:
            return False

    return True


def _revenue_bounds(text: str) -> tuple[Optional[float], Optional[float]]:
    """Return the numeric bounds described by a revenue range string in millions."""

    if text is None:
        return (None, None)

    cleaned = str(text).replace(",", "").upper().strip()

    m = re.match(r"^<\$?\s*(\d+(?:\.\d+)?)\s*([MB])$", cleaned)
    if m:
        val = float(m.group(1))
        if m.group(2) == "B":
            val *= 1000
        return (None, val)

    m = re.match(r"^>\$?\s*(\d+(?:\.\d+)?)\s*([MB])$", cleaned)
    if m:
        val = float(m.group(1))
        if m.group(2) == "B":
            val *= 1000
        return (val, None)

    m = re.match(r"^\$?\s*(\d+(?:\.\d+)?)\s*([MB])\s*(?:TO|[-–])\s*\$?\s*(\d+(?:\.\d+)?)\s*([MB])$", cleaned)
    if m:
        low = float(m.group(1))
        if m.group(2) == "B":
            low *= 1000
        high = float(m.group(3))
        if m.group(4) == "B":
            high *= 1000
        return (low, high)

    m = re.match(r"^\$?\s*(\d+(?:\.\d+)?)\s*([MB])\+?$", cleaned)
    if m:
        val = float(m.group(1))
        if m.group(2) == "B":
            val *= 1000
        return (val, None)

    return (None, None)


def _revenue_range_contains(default: str, actual: str) -> bool:
    """Return True if ``actual`` falls within ``default`` revenue bounds."""

    d_min, d_max = _revenue_bounds(default)
    a_min, a_max = _revenue_bounds(actual)

    if d_min is None or a_min is None:
        return False

    if a_min < d_min:
        return False

    if d_max is not None:
        if a_max is None or a_max > d_max:
            return False

    return True


def _parse_industries(value: Optional[str]) -> Optional[list[str]]:
    """Split the raw ``Industries`` string into a list of values."""

    if value is None:
        return None
    parts = [p.strip() for p in re.split(r"[;,]", value) if p.strip()]
    return parts or None


def _is_business(name: str) -> bool:
    """Return True if the name appears to describe a business."""
    text = name.lower()
    return not any(re.search(pat, text) for pat in _EXCLUDE_PATTERNS)


def read_companies_from_csv(path: Union[str, Path]) -> List[Company]:
    """Read company data from a CSV file and return a list of Company objects."""

    path = Path(path)
    companies: List[Company] = []

    with path.open("rb") as raw:
        sample_bytes = raw.read(2048)
        raw.seek(0)
        try:
            sample = sample_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            sample = sample_bytes.decode("latin-1").encode("utf-8", "replace").decode("utf-8")

        try:
            dialect = csv.Sniffer().sniff(sample)
            if dialect.delimiter not in {",", ";", "\t", "|"}:
                raise csv.Error("unlikely delimiter")
        except csv.Error:
            dialect = csv.excel

        reader = csv.DictReader(_decode_lines(raw), dialect=dialect)

        sanitized_map = {_sanitize(v): k for k, v in CANONICAL_HEADERS.items()}
        field_map: Dict[str, str] = {}
        for header in reader.fieldnames or []:
            key = sanitized_map.get(_sanitize(header))
            if key:
                field_map[key] = header

        for row in reader:
            kwargs = {}
            for attr in CANONICAL_HEADERS.keys():
                header = field_map.get(attr)
                value = row.get(header) if header else None
                if value is not None:
                    value = value.strip()
                if attr == "organization_name":
                    kwargs[attr] = value or ""
                elif attr == "number_of_employees" and value is not None:
                    kwargs[attr] = _fix_employee_range(value) if value != "" else None
                elif attr == "industries" and value is not None:
                    kwargs[attr] = _parse_industries(value)
                else:
                    kwargs[attr] = value if value != "" else None

            company = Company(**kwargs)
            if not _is_business(company.organization_name):
                continue
            companies.append(company)

    return companies


def read_companies_from_xlsx(path: Union[str, Path]) -> List[Company]:
    """Read company data from an XLSX file and return a list of ``Company`` objects.

    This function requires the optional :mod:`openpyxl` package. If it is not
    installed, an :class:`ImportError` is raised with a helpful message.
    """

    try:
        import openpyxl  # type: ignore
    except Exception as exc:  # pragma: no cover - optional dependency
        raise ImportError(
            "openpyxl is required to read .xlsx files"
        ) from exc

    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    header_cells = [str(c.value or "") for c in next(ws.iter_rows(min_row=1, max_row=1))]

    sanitized_map = {_sanitize(v): k for k, v in CANONICAL_HEADERS.items()}
    field_map: Dict[str, int] = {}
    for idx, header in enumerate(header_cells):
        key = sanitized_map.get(_sanitize(header))
        if key:
            field_map[key] = idx

    companies: List[Company] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        kwargs = {}
        for attr in CANONICAL_HEADERS.keys():
            idx = field_map.get(attr)
            value = row[idx] if idx is not None and idx < len(row) else None
            if isinstance(value, str):
                value = value.strip()
            if attr == "organization_name":
                kwargs[attr] = value or ""
            elif attr == "industries" and value not in (None, ""):
                kwargs[attr] = _parse_industries(str(value))
            else:
                kwargs[attr] = value if value not in (None, "") else None

        company = Company(**kwargs)
        if not _is_business(company.organization_name):
            continue
        companies.append(company)

    return companies


def _parse_filename_metadata(path: Path) -> Dict[str, Optional[str]]:
    """Return field defaults extracted from a Crunchbase search filename."""

    stem = path.stem
    if stem.lower().startswith("fs_"):
        stem = stem[3:]
    parts = stem.split("_")

    operating_status: Optional[str] = None
    ipo_status: Optional[str] = None
    revenue: Optional[str] = None
    employees: Optional[str] = None

    if len(parts) >= 4:
        operating_status, ipo_status, revenue, employees = parts[:4]
    elif len(parts) == 3:
        token = parts[0]
        m = re.match(r"([A-Za-z]+?)([A-Z].*)$", token)
        if m:
            operating_status, ipo_status = m.groups()
        else:
            operating_status = token
        revenue = parts[1]
        employees = parts[2]
    else:
        return {}

    def _format_rev(token: str) -> str:
        m = re.match(r"(?i)(\d+(?:\.\d+)?)([MB])to(\d+(?:\.\d+)?)([MB])", token)
        if m:
            return f"${m.group(1)}{m.group(2).upper()} to ${m.group(3)}{m.group(4).upper()}"
        m = re.match(r"(?i)(\d+(?:\.\d+)?)([MB])plus", token)
        if m:
            return f">${m.group(1)}{m.group(2).upper()}"
        return token

    def _format_emp(token: str) -> str:
        m = re.match(r"(?i)(\d+)to(\d+)", token)
        if m:
            return f"{m.group(1)}-{m.group(2)}"
        m = re.match(r"(?i)(\d+)plus", token)
        if m:
            return f"{m.group(1)}+"
        return token

    return {
        "operating_status": operating_status,
        "ipo_status": ipo_status,
        "estimated_revenue_range": _format_rev(revenue) if revenue else None,
        "number_of_employees": _format_emp(employees) if employees else None,
    }


def read_companies_from_csvs(paths: Iterable[Union[str, Path]]) -> List[Company]:
    """Read companies from multiple CSV files, applying filename metadata."""

    all_companies: List[Company] = []
    for path in paths:
        p = Path(path)
        defaults = _parse_filename_metadata(p)
        companies = read_companies_from_csv(p)
        for idx, company in enumerate(companies):
            for field, value in defaults.items():
                if value is None:
                    continue
                current = getattr(company, field)
                if current in (None, ""):
                    setattr(company, field, value)
                elif str(current).strip().lower() != str(value).strip().lower():
                    if field == "number_of_employees" and _employee_range_contains(value, str(current)):
                        continue
                    if field == "estimated_revenue_range" and _revenue_range_contains(value, str(current)):
                        continue
                    logger.warning(
                        "%s row %d %s %r conflicts with filename %r",
                        p.name,
                        idx,
                        field,
                        current,
                        value,
                    )
        all_companies.extend(companies)

    return all_companies
